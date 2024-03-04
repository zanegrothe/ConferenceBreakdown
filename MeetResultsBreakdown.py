import requests
from bs4 import BeautifulSoup
import lxml
import openpyxl as xl
import pandas as pd

# 2024 SEC
# https://sidearmstats.com/auburn/swim/
# 2023 SEC
# https://sidearmstats.com/tamu/swim/
# 2024 W ACC
# https://sidearmstats.com/acc/swimming/
# 2024 W B1G
# https://sidearmstats.com/purdue/swimming/


y = 0
loops = 1  # Men's meet or Women's meet (increase to 2 if combined)
while y < loops:
    if y == 0:
        print("Enter this year's live results url and press enter:")
        meet_results_main = input("> ")
    # else:
    #     print("Enter last year's live results url and press enter")
    #     meet_results_main = input("> ")

    # Create soup
    index_site = meet_results_main + 'evtindex.htm'
    index_text = requests.get(index_site).text
    index_soup = BeautifulSoup(index_text, 'lxml')

    # Determine if combined
    combined = index_soup.find('h2').text
    if 'Women' or 'Men' in combined:
        loops = 1
    else:
        loops = 2
    print(combined)

    # # Find the last event
    # events = index_soup.find_all('a')
    # if events[-1].text[2] == ' ':
    #     d = 2
    #     total_events_code = '0' * d + events[-1].text[1]
    # elif events[-1].text[3] == ' ':
    #     d = 1
    #     total_events_code = '0' * d + events[-1].text[1] + events[-1].text[2]
    # elif events[-1].text[4] == ' ':
    #     d = 0
    #     total_events_code = '0' * d + events[-1].text[1] + events[-1].text[2] + events[-1].text[3]
    #
    # # Create date code
    # date = index_soup.find('p')
    # if int(date.text[0]) > 1:
    #     day = date.text[2] + date.text[3]
    #     month = str(0) + date.text[0]
    #     year = date.text[7] + date.text[8]
    # else:
    #     day = date.text[3] + date.text[4]
    #     month = date.text[0] + date.text[1]
    #     year = date.text[8] + date.text[9]
    # date_code = year + month + day
    # if y == 0:
    #     current_year = year
    #
    # # Find the final scores  ###################################################################
    # if y == 0:
    #     final_event_code = total_events_code
    # else:
    #     final_event_code = '0' * d + str(int(total_events_code) - 1)
    #
    #
    # final_scores_page = meet_results_main + date_code + 'F' + final_event_code + '.htm'
    # fsp_text = requests.get(final_scores_page).text
    # fsp_soup = BeautifulSoup(fsp_text, 'lxml')
    # fsp_pre = fsp_soup.find('pre').text
    # team_rankings_start = fsp_pre.find('Team Rankings')
    #
    #
    # print(team_rankings_start)

    # men_team_rankings_text = fsp_pre_M[men_team_rankings_start:]
    #
    # men_ranking_list = men_team_rankings_text.split()
    #
    # # Index ranking places (1., 2., ...)
    # b = True
    # k = 1
    # mrl_indices = []
    # while b:
    #     try:
    #         mrl_indices.append(men_ranking_list.index(f'{k}.'))
    #     except ValueError:
    #         b = False
    #     k += 1
    #
    # # List teams in order of rankings
    # k = 0
    # team_list_M = []
    # while k < len(mrl_indices):
    #     if k < len(mrl_indices) - 1:
    #         team_list_M.append(' '.join(men_ranking_list[mrl_indices[k] + 1: mrl_indices[k + 1] - 1]))
    #     else:
    #         team_list_M.append(' '.join(men_ranking_list[mrl_indices[k] + 1: -1]))
    #     k += 1
    #
    # # List team scores in order of rankings
    # numbers_in_rankings_M = []
    # for t in men_team_rankings_text.split():
    #     try:
    #         numbers_in_rankings_M.append(float(t))
    #     except ValueError:
    #         pass
    # scores_M = numbers_in_rankings_M[2::2]
    y += 1
    # print(team_list_M)
#
#     teams = ['first', 'second', 'third', 'fourth', 'fifth',
#              'sixth', 'seventh', 'eighth', 'ninth', 'tenth',
#              'eleventh', 'twelfth', 'thirteenth', 'fourteenth', 'fifteenth']
#     k = 1
#     while k < len(scores_M) + 2:
#         teams[k - 1] = {
#             "rank": k,
#             "team": team1_M,
#             "score": scores_M[k-1]
#         }
#
#     # Women
#     women_final_event_code = '0' + str(int(total_events_code) - 1)
#     final_scores_page_W = meet_results_main + date_code + 'F' + women_final_event_code + '.htm'
#     fsp_text_W = requests.get(final_scores_page_W).text
#     fsp_soup_W = BeautifulSoup(fsp_text_W, 'lxml')
#     fsp_pre_W = fsp_soup_W.find('pre').text
#     women_team_rankings_start = fsp_pre_W.find('Team Rankings')
#     women_team_rankings_text = fsp_pre_W[women_team_rankings_start:]
#
#     women_ranking_list = women_team_rankings_text.split()
#     first_W = women_team_rankings_text.split().index('1.')
#     second_W = women_team_rankings_text.split().index('2.')
#     third_W = women_team_rankings_text.split().index('3.')
#     fourth_W = women_team_rankings_text.split().index('4.')
#     fifth_W = women_team_rankings_text.split().index('5.')
#     sixth_W = women_team_rankings_text.split().index('6.')
#     seventh_W = women_team_rankings_text.split().index('7.')
#     eighth_W = women_team_rankings_text.split().index('8.')
#     ninth_W = women_team_rankings_text.split().index('9.')
#     tenth_W = women_team_rankings_text.split().index('10.')
#     eleventh_W = women_team_rankings_text.split().index('11.')
#     twelfth_W = women_team_rankings_text.split().index('12.')
#     team1_W = ' '.join(women_ranking_list[first_W + 1: second_W - 1])
#     team2_W = ' '.join(women_ranking_list[second_W + 1: third_W - 1])
#     team3_W = ' '.join(women_ranking_list[third_W + 1: fourth_W - 1])
#     team4_W = ' '.join(women_ranking_list[fourth_W + 1: fifth_W - 1])
#     team5_W = ' '.join(women_ranking_list[fifth_W + 1: sixth_W - 1])
#     team6_W = ' '.join(women_ranking_list[sixth_W + 1: seventh_W - 1])
#     team7_W = ' '.join(women_ranking_list[seventh_W + 1: eighth_W - 1])
#     team8_W = ' '.join(women_ranking_list[eighth_W + 1: ninth_W - 1])
#     team9_W = ' '.join(women_ranking_list[ninth_W + 1: tenth_W - 1])
#     team10_W = ' '.join(women_ranking_list[tenth_W + 1: eleventh_W - 1])
#     team11_W = ' '.join(women_ranking_list[eleventh_W + 1: twelfth_W - 1])
#     team12_W = ' '.join(women_ranking_list[twelfth_W + 1: -1])
#
#     numbers_in_rankings = []
#     for t in women_team_rankings_text.split():
#         try:
#             numbers_in_rankings.append(float(t))
#         except ValueError:
#             pass
#     scores_W = numbers_in_rankings[2::2]
#
#     y += 1
#
# # Write data to an Excel workbook ####################################################
# workbook = xl.Workbook()
#
# # Select the default sheet (usually named 'Sheet')
# sheet = workbook.active
#
# sheet.append([f"20{current_year} SEC Team Scores"])
#
# # Add Men's data to the Excel sheet
# rankings_M = [
#     ["Men's Teams"],
#     ["", "School", "Points"],
#     [1, team1_M, scores_M[0]],
#     [2, team2_M, scores_M[1]],
#     [3, team3_M, scores_M[2]],
#     [4, team4_M, scores_M[3]],
#     [5, team5_M, scores_M[4]],
#     [6, team6_M, scores_M[5]],
#     [7, team7_M, scores_M[6]],
#     [8, team8_M, scores_M[7]],
#     [9, team9_M, scores_M[8]],
#     [10, team10_M, scores_M[9]]
# ]
# for row in rankings_M:
#     sheet.append(row)
#
# sheet.cell(row=14, column=3).value = '=sum(C4:C13)'
#
# # Add Women's data to the Excel sheet
# rankings_W = [
#     ["Women's Teams"],
#     ["", "School", "Points"],
#     [1, team1_W, scores_W[0]],
#     [2, team2_W, scores_W[1]],
#     [3, team3_W, scores_W[2]],
#     [4, team4_W, scores_W[3]],
#     [5, team5_W, scores_W[4]],
#     [6, team6_W, scores_W[5]],
#     [7, team7_W, scores_W[6]],
#     [8, team8_W, scores_W[7]],
#     [9, team9_W, scores_W[8]],
#     [10, team10_W, scores_W[9]],
#     [11, team11_W, scores_W[10]],
#     [12, team12_W, scores_W[11]]
# ]
# for row in rankings_W:
#     sheet.append(row)
#
# sheet.cell(row=29, column=3).value = '=sum(C17:C28)'
#
# # Adjust dimensions
# sheet.column_dimensions['A'].width = 3
# sheet.column_dimensions['B'].width = 30
# sheet.column_dimensions['C'].width = 8
#
# # Merge Cells
# sheet.merge_cells('A1:C1')
# sheet.merge_cells('A2:B2')
# sheet.merge_cells('A15:B15')
#
# # Save the workbook to a file
# workbook.save(f"20{current_year}SECtest.xlsx")
#
# # Print a success message
# print("Excel file created successfully!")
